from django.http import HttpResponse
from django.shortcuts import redirect
from openpyxl import Workbook
from .models import Cdp
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill
from ..establecimiento.models import Establecimiento
from .models_proyeccion import MesProyectado, Subvencion, TIPO_MONTO
from ..general.templatetags.utilidades import montoConPuntos
from django.shortcuts import get_object_or_404
from datetime import datetime
from django.contrib import messages
from django.db.models import Sum


def exportar_cdps(request, year, program, establecimiento):
    cdps = Cdp.objects.all()
    nombre_archivo = "Reporte_cdp"
    if year != 0:
        cdps = cdps.filter(item_presupuestario__subtitulo_presupuestario__year__year=year)
        nombre_archivo += "_" + str(year)

    if program != 'todos':
        cdps = cdps.filter(item_presupuestario__subtitulo_presupuestario__programa_presupuestario=program)
        nombre_archivo += "_" + program

    if establecimiento != 'todos':
        cdps = cdps.filter(establecimiento=establecimiento)
        nombre_establecimiento = get_object_or_404(Establecimiento, id=establecimiento).nombre
        nombre_archivo += "_" + nombre_establecimiento

    # Crear un libro de trabajo y una hoja
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'CDPs'

    # Definir los encabezados
    headers = ['FECHA CDP', 'FECHA GUIA REQUERIMIENTO', 'N° REQUERIMIENTO', 'RBD', 'ESTABLECIMIENTO', 'PROGRAMA', 'FONDO', 'CUENTA CONTABLE', 'CDP', 'FOLIO SIGFE', 'N° DE ORDEN', 'MONTO', 'DETALLES', 'OTROS']
    worksheet.append(headers)

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Aplicar formato a los encabezados
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True)
    header_style.alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.style = header_style
        cell.border = thin_border  # Aplicar bordes

    # Estilo para centrado
    centered_style = NamedStyle(name="centered")
    centered_style.alignment = Alignment(horizontal="center", vertical="center")

    cdps = cdps.order_by('cdp')
    
    # Agregar los datos a la hoja
    for cdp in cdps:
        rbd = ""
        establecimiento = ""
        if cdp.establecimiento:
            rbd = cdp.establecimiento.rbd
            establecimiento = cdp.establecimiento.nombre
        else:
            establecimiento = "SERVICIO LOCAL PUERTO CORDILLERA"
        programa = ""
        if cdp.item_presupuestario.subtitulo_presupuestario.programa_presupuestario == "P01 GASTOS ADMINISTRATIVOS":
            programa = "01"
        else:
            programa = "02"

        otros = cdp.otros if cdp.otros else "-"  # Usar una condición para manejar celdas vacías o None



        worksheet.append([
            cdp.fecha_cdp.strftime('%d/%m/%Y'),  # Listo
            cdp.fecha_guia_requerimiento.strftime('%d/%m/%Y'),  # Listo
            cdp.numero_requerimiento,  # Listo
            rbd,  # Listo
            establecimiento,  # Listo
            programa,  # Listo
            cdp.fondo,  # Listo
            cdp.item_presupuestario.concepto_presupuestario_item,  # Listo
            int(cdp.cdp),  # Listo
            cdp.folio_sigfe,  # Listo
            cdp.n_orden,  # Listo
            int(cdp.monto),  # Listo
            cdp.detalle,  # Listo
            otros,  # Listo
        ])

    # Aplicar formato a las celdas
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for col_idx, cell in enumerate(row, start=1):
            cell.border = thin_border  # Aplicar bordes negros a cada celda
            cell.alignment = centered_style.alignment

            if col_idx == 12:  # Columna del monto
                cell.number_format = '#,##0'
                
            elif col_idx == 8 or col_idx == 13 or col_idx == 14:  # Columna del concepto presupuestario
                pass
            else:
                cell.alignment = Alignment(wrap_text=True)
            
            #if cell.row % 2 != 0:
            #        cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    
    for i in range(2, worksheet.max_row + 1):#Iterar sobre las filas
        for j in range(1 , worksheet.max_column + 1):
            if i % 2 == 0:
                cell = worksheet.cell(row=i, column=j)
                cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    # Ajustar el ancho de las columnas según el contenido, con límite a 50
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Obtener la letra de la columna
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                    if max_length > 50:
                        max_length = 50
                else:
                    max_length = 10  # Si es vacío, limitar el ancho
            except:
                pass
        adjusted_width = max_length + 10  # Añadir un poco de espacio extra
        worksheet.column_dimensions[column].width = adjusted_width

    # Establecer un ancho fijo para la columna 13 (letra M y N), asegurando que no se sobrepongan
    worksheet.column_dimensions['M'].width = 50
    worksheet.column_dimensions['N'].width = 20  # Columna N más estrecha para evitar solapamiento

    

    fecha_actual = datetime.now().strftime('%d/%m/%Y')
    nombre_archivo += "__" + fecha_actual
    
    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}.xlsx'
    workbook.save(response)
    return response


def exportar_meses_proyectados(request, establecimiento, subvencion, tipo):
    print(f"Establecimiento: {establecimiento} Subvencion: {subvencion} tipo: {tipo}")
    establecimiento = Establecimiento.objects.get(id=establecimiento)
    subvencion = Subvencion.objects.get(id=subvencion)

    nombre_archivo = f"meses_proyectados_{establecimiento.rbd}_{subvencion.fondo}_{tipo}"
    # Crear un libro de trabajo y una hoja
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'meses_proyectados'

    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Crear las cabeceras con saltos de línea
    cabeceras = []
    for mes in MesProyectado.objects.filter(subvencion=subvencion, tipo=tipo):
        cabecera = f"{mes.mes}\n{mes.estado}"
        cabeceras.append(cabecera)

    # Definir los encabezados
    headers = ['RBD', 'ESTABLECIMIENTO'] + cabeceras + ['TOTAL']
    date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    year = datetime.now().year

    # Añadir una fila superior con el título "Ingresos Subvención Normal 2024"
    if tipo == TIPO_MONTO[0][0]:  # Ingreso
        titulo = f"Estimación Ingresos - {subvencion.fondo} - {year}"
    elif tipo == TIPO_MONTO[1][0]:  # Remuneraciones
        titulo = f"Estimación Remuneraciones - {subvencion.fondo} - {year}"
    
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers)-2)  # Dividir el título y la fecha
    titulo_cell = worksheet.cell(row=1, column=1)
    titulo_cell.value = titulo
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    titulo_cell.font = Font(size=14, bold=True)  # Estilizar el texto del título
    titulo_cell.border = thin_border
    
    
    # Añadir la fecha y hora en la celda superior derecha
    worksheet.merge_cells(start_row=1, start_column=len(headers)-1, end_row=1, end_column=len(headers))  # Segunda celda para la fecha
    fecha_hora_cell = worksheet.cell(row=1, column=len(headers)-1)
    fecha_hora_cell.value = f"Generado el: {date}"
    fecha_hora_cell.alignment = Alignment(horizontal='right', vertical='center')
    fecha_hora_cell.font = Font(size=10, italic=True)
    fecha_hora_cell.border = thin_border
    worksheet.row_dimensions[1].border = thin_border
    # Añadir las cabeceras en la fila 2
    worksheet.append(headers)

    # Aplicar ajuste de texto, centrar las cabeceras y ajustar el tamaño de las celdas
    for col_number, header in enumerate(headers, start=1):  # Iterar sobre las columnas y cabeceras
        cell = worksheet.cell(row=2, column=col_number)  # Segunda fila
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # Ajustar y centrar
        cell.border = thin_border  # Agregar bordes
        
        # Ajustar el ancho de la columna en función de la longitud del texto
        worksheet.column_dimensions[cell.column_letter].width = len(header)  # Añade un margen para mejor visualización
    # Ajustar específicamente el ancho de la columna A
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['O'].width = 15

    # Crear la fila con los datos del establecimiento
    fila = [establecimiento.rbd, establecimiento.nombre]
    for mes in MesProyectado.objects.filter(subvencion=subvencion, tipo=tipo):
        fila.append(mes.monto)
    if tipo == TIPO_MONTO[0][0]:  # Ingreso
        fila.append(subvencion.monto_total_proyectado_ingreso)
    elif tipo == TIPO_MONTO[1][0]:  # Remuneraciones
        fila.append(subvencion.monto_total_proyectado_remuneraciones)
    worksheet.append(fila)

    # Aplicar formato numérico a las celdas de montos
    row_number = worksheet.max_row  # Última fila añadida
    for col_number in range(3, len(fila) + 1):  # Desde la tercera columna
        cell = worksheet.cell(row=row_number, column=col_number)
        cell.number_format = '#,##0'  # Formato numérico
    
    for col_number in range(1, len(fila) + 1):  # Desde la primera columna
        cell = worksheet.cell(row=row_number, column=col_number)  # Asociar celda específica
        cell.border = thin_border  # Agregar bordes

    for col_numer in range (1, len(headers)+1):
        cell = worksheet.cell(row=2, column=col_numer)
        fill_color = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
        cell.fill = fill_color
    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}.xlsx'
    workbook.save(response)

    return response




def exportar_estimacion_subvencion(request, establecimiento, subvencion):
    try:
        establecimiento = Establecimiento.objects.get(id=establecimiento)
        subvencion = Subvencion.objects.get(id=subvencion)
    except:
        messages.info(request, "Surgio un error inesperado")
        return redirect('ingresar_proyeccion_inicial','todos')

    nombre_archivo = "Estimacion " + subvencion.fondo
    # Crear un libro de trabajo y una hoja
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'CDPs'

    # Definir los encabezados
    headers = ['RBD', 'ESTABLECIMIENTO', 'PROYECCION\nINGRESOS', 'REMUNERACIONES\nPROYECTADAS', 'SALDO DISPONIBLE\nPARA GASTOS\nOPERATIVOS', 'CDP EMITIDOS', 'SALDO', '% REMUNERACIONES', '% GASTOS OPERATIVOS']
    

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )
    date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))  # Dividir el título y la fecha
    titulo_cell = worksheet.cell(row=1, column=1)
    titulo_cell.value = f"ESTIMACIÓN SUBVENCION {subvencion.fondo.upper()}- {date}"
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    titulo_cell.font = Font(size=14, bold=True)  # Estilizar el texto del título
    titulo_cell.border = thin_border
    worksheet.append(headers)

    # Aplicar bordes a todas las celdas dentro del rango combinado
    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Formulas para calcular los valores de las casillas

    saldo_disponible_operativos = subvencion.monto_total_proyectado_ingreso - subvencion.monto_total_proyectado_remuneraciones
    monto_cdp = 0
    for cdp in Cdp.objects.filter(establecimiento=establecimiento,fondo=subvencion.fondo):
        monto_cdp += cdp.monto 
    saldo = saldo_disponible_operativos - monto_cdp

    remuneraciones = f"{round(100*(subvencion.monto_total_proyectado_remuneraciones/subvencion.monto_total_proyectado_ingreso))}%"
    gastos_normativos = f"{round(100*((subvencion.monto_total_proyectado_remuneraciones+monto_cdp)/subvencion.monto_total_proyectado_ingreso))}%"
    worksheet.append([
            establecimiento.rbd,
            establecimiento.nombre,
            subvencion.monto_total_proyectado_ingreso,
            subvencion.monto_total_proyectado_remuneraciones,
            saldo_disponible_operativos,
            monto_cdp,
            saldo,
            remuneraciones,
            gastos_normativos,
        ])

    row_number = worksheet.max_row  # Última fila añadida
    # Ajustar formato a los headers (fila 2)
    for col_number, header in enumerate(headers, start=1):  # Iterar sobre las columnas y cabeceras
        cell = worksheet.cell(row=2, column=col_number)  # Segunda fila
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # Ajustar y centrar
        cell.font = Font(bold=True)  # Negrita para los headers
        cell.border = thin_border  # Agregar bordes a los headers
        cell2 = worksheet.cell(row=3, column=col_number)
        cell2.border = thin_border  # Agregar bordes a los valores
        # Dividir el texto del header por el salto de línea
        max_length = max(len(part) for part in header.split('\n'))
        # Ajustar el ancho de la columna según la longitud máxima
        column_letter = worksheet.cell(row=2, column=col_number).column_letter
        worksheet.column_dimensions[column_letter].width = max_length + 4  # Ajuste extra para margen

        cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')


    for col_number in range (3, len(headers)-1):
        cell = worksheet.cell(row=3, column=col_number)
        cell.number_format = '#,##0'

    worksheet.column_dimensions['B'].width = 20

    # Aplicar formato numérico a las celdas de montos
    cell_value = worksheet.cell(row=3, column=7).value
    text_length = len(str(cell_value))
    print(text_length)

    worksheet.column_dimensions['G'].width = text_length + 4

    cell_value = worksheet.cell(row=2, column=8).value
    text_length = len(cell_value)
    worksheet.column_dimensions['H'].width = text_length + 4

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}.xlsx'
    workbook.save(response)
    return response